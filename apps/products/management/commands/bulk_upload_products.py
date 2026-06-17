import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.products.models import Brand, Category, Product


class Command(BaseCommand):
    help = (
        "Bulk import products, brands, and categories from an Excel file."
        " Categories and brands are created only when they do not already exist."
    )

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Path to the Excel file.")
        parser.add_argument(
            "--sheet",
            type=str,
            default=None,
            help="Optional sheet name. Uses the first sheet when omitted.",
        )
        parser.add_argument(
            "--skip-products",
            action="store_true",
            help="Only import brands/categories, do not create products.",
        )
        parser.add_argument(
            "--skip-brands-categories",
            action="store_true",
            help="Only import products; assume brands/categories already exist.",
        )

    def handle(self, *args, **options):
        path = Path(options["xlsx_path"])
        if not path.exists():
            raise CommandError(f"Excel file not found: {path}")

        workbook = load_workbook(path, data_only=True)
        sheet_name = options["sheet"]
        worksheet = workbook[sheet_name] if sheet_name else workbook.active

        rows = list(worksheet.iter_rows(values_only=True))
        if len(rows) < 2:
            raise CommandError(
                "Excel file must contain a header row and at least one data row."
            )

        headers = [self.normalize_header(cell) for cell in rows[0]]
        if not any(headers):
            raise CommandError(
                "Failed to parse any header columns from the Excel file."
            )

        self.stdout.write(self.style.NOTICE(f"Importing sheet: {worksheet.title}"))
        self.stdout.write(self.style.NOTICE(f"Columns detected: {headers}"))

        sale_price_headers = {
            "sale price",
            "selling price",
            "selling_price",
            "sale_price",
            "discounted price",
            "discounted_price",
            "offer price",
            "offer_price",
            "special price",
            "special_price",
        }
        discount_headers = {
            "discount",
            "discount amount",
            "discount_amount",
            "discount_value",
        }
        discount_column_is_sale_price = False
        if not sale_price_headers.intersection(headers):
            sample_rows = rows[1 : min(len(rows), 100)]
            sample_values = []
            for row in sample_rows:
                row_values = {
                    headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))
                }
                sample_price = self.parse_nullable_decimal(
                    self.get_any(
                        row_values,
                        [
                            "price",
                            "mrp",
                            "selling price",
                            "selling_price",
                            "sale_price",
                        ],
                    )
                )
                sample_discount = self.parse_nullable_decimal(
                    self.get_any(row_values, list(discount_headers))
                )
                if (
                    sample_price is not None
                    and sample_price > 0
                    and sample_discount is not None
                ):
                    sample_values.append((sample_price, sample_discount))
            discount_column_is_sale_price = self.detect_discount_column_as_sale_price(
                sample_values
            )
            if discount_column_is_sale_price:
                self.stdout.write(
                    self.style.NOTICE(
                        "Detected discount_amount column as sale price values; "
                        "computing discount_amount as price - sale_price."
                    )
                )

        summary = {
            "brands_created": 0,
            "categories_created": 0,
            "subcategories_created": 0,
            "products_created": 0,
            "products_skipped": 0,
            "rows_processed": 0,
        }

        with transaction.atomic():
            for row_number, row in enumerate(rows[1:], start=2):
                values = {
                    headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))
                }
                summary["rows_processed"] += 1

                brand_name = self.normalize_value(
                    self.get_any(
                        values, ["brand", "brand name", "manufacturer", "vendor"]
                    )
                )
                category_name = self.normalize_value(
                    self.get_any(values, ["category", "main category", "main_category"])
                )
                subcategory_name = self.normalize_value(
                    self.get_any(
                        values, ["subcategory", "sub category", "sub_category"]
                    )
                )

                product_category = None
                if not options["skip_brands_categories"]:
                    if category_name:
                        category, created = Category.objects.get_or_create(
                            name=category_name
                        )
                        if created:
                            summary["categories_created"] += 1
                        product_category = category
                    if subcategory_name:
                        subcategory, created = Category.objects.get_or_create(
                            name=subcategory_name
                        )
                        if created:
                            summary["subcategories_created"] += 1
                        product_category = subcategory

                    if brand_name:
                        brand, created = Brand.objects.get_or_create(name=brand_name)
                        if created:
                            summary["brands_created"] += 1
                    else:
                        brand = None
                else:
                    brand = None
                    if subcategory_name:
                        product_category = Category.objects.filter(
                            name=subcategory_name
                        ).first()
                    elif category_name:
                        product_category = Category.objects.filter(
                            name=category_name
                        ).first()

                if options["skip_products"]:
                    continue

                product_name = (
                    self.normalize_value(
                        self.get_any(
                            values,
                            [
                                "name",
                                "product name",
                                "item name",
                                "product",
                                "item",
                            ],
                        )
                    )
                    or f"Untitled Product {row_number}"
                )
                sku = self.normalize_value(
                    self.get_any(values, ["sku", "sku code", "sku_no", "sku_number"])
                )
                if not sku:
                    sku = f"SKU-{timezone.now().strftime('%Y%m%d%H%M%S')}-{row_number}"

                if Product.objects.filter(sku=sku).exists():
                    summary["products_skipped"] += 1
                    self.stdout.write(
                        self.style.WARNING(
                            f"Row {row_number}: SKU already exists, skipping product '{product_name}' (SKU={sku})."
                        )
                    )
                    continue

                price = self.parse_decimal(
                    self.get_any(
                        values,
                        [
                            "price",
                            "mrp",
                            "selling price",
                            "selling_price",
                            "sale_price",
                        ],
                    )
                )
                raw_discount = self.parse_decimal(
                    self.get_any(
                        values,
                        [
                            "discount",
                            "discount amount",
                            "discount_amount",
                            "discount_value",
                        ],
                    )
                )
                sale_price = self.parse_nullable_decimal(
                    self.get_any(
                        values,
                        [
                            "sale price",
                            "sale_price",
                            "selling price",
                            "selling_price",
                            "discounted price",
                            "discounted_price",
                            "offer price",
                            "offer_price",
                            "special price",
                            "special_price",
                        ],
                    )
                )
                discount_amount = self.compute_discount_amount(
                    price,
                    raw_discount,
                    sale_price,
                    discount_column_is_sale_price,
                )
                unit = (
                    self.normalize_value(
                        self.get_any(
                            values,
                            [
                                "unit",
                                "quantity unit",
                                "qty_unit",
                                "measure",
                                "measure unit",
                            ],
                        )
                    )
                    or "per unit"
                )
                description = self.normalize_value(
                    self.get_any(
                        values,
                        [
                            "description",
                            "details",
                            "product description",
                            "product_description",
                        ],
                    )
                )
                in_stock = self.parse_bool(
                    self.get_any(
                        values,
                        [
                            "in_stock",
                            "in stock",
                            "stock",
                            "stock_available",
                            "available",
                        ],
                    )
                )
                key_detail_title = self.normalize_value(
                    self.get_any(values, ["key_detail_title", "key detail title"])
                )
                key_detail_description = self.normalize_value(
                    self.get_any(
                        values, ["key_detail_description", "key detail description"]
                    )
                )
                is_hot_deal = self.parse_bool(
                    self.get_any(values, ["is_hot_deal", "hot deal", "hot_deal"])
                )
                hot_deal_start = self.parse_datetime(
                    self.get_any(
                        values, ["hot_deal_start", "hot_deal_begin", "deal_start"]
                    )
                )
                hot_deal_end = self.parse_datetime(
                    self.get_any(
                        values, ["hot_deal_end", "hot_deal_end_date", "deal_end"]
                    )
                )

                product = Product.objects.create(
                    brand=brand,
                    category=product_category,
                    name=product_name,
                    description=description or "",
                    sku=sku,
                    price=price,
                    discount_amount=discount_amount,
                    unit=unit,
                    in_stock=in_stock,
                    key_detail_title=key_detail_title,
                    key_detail_description=key_detail_description,
                    is_hot_deal=is_hot_deal,
                    hot_deal_start=hot_deal_start,
                    hot_deal_end=hot_deal_end,
                )
                summary["products_created"] += 1

        self.stdout.write(self.style.SUCCESS("Import complete."))
        self.stdout.write(
            self.style.SUCCESS(f"Brands created: {summary['brands_created']}")
        )
        self.stdout.write(
            self.style.SUCCESS(f"Categories created: {summary['categories_created']}")
        )
        self.stdout.write(
            self.style.SUCCESS(
                f"Subcategories created: {summary['subcategories_created']}"
            )
        )
        self.stdout.write(
            self.style.SUCCESS(f"Products created: {summary['products_created']}")
        )
        self.stdout.write(
            self.style.WARNING(f"Products skipped: {summary['products_skipped']}")
        )

    @staticmethod
    def normalize_header(value):
        if value is None:
            return ""
        normalized = str(value).strip().lower()
        normalized = re.sub(r"[\s\-_]+", " ", normalized).strip()
        return normalized

    @staticmethod
    def normalize_value(value):
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return value

    @staticmethod
    def get_any(values, keys):
        for key in keys:
            normalized_key = Command.normalize_header(key)
            if normalized_key in values and values[normalized_key] is not None:
                return values[normalized_key]
        return None

    @staticmethod
    def parse_decimal(value):
        if value is None or value == "":
            return Decimal("0.00")
        if isinstance(value, Decimal):
            return value
        try:
            return Decimal(str(value)).quantize(Decimal("0.00"))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal("0.00")

    @staticmethod
    def parse_nullable_decimal(value):
        if value is None or value == "":
            return None
        if isinstance(value, Decimal):
            return value
        try:
            return Decimal(str(value)).quantize(Decimal("0.00"))
        except (InvalidOperation, TypeError, ValueError):
            return None

    @staticmethod
    def parse_bool(value):
        if value is None or value == "":
            return True
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        return text not in {"false", "0", "no", "n", "none"}

    @staticmethod
    def compute_discount_amount(
        price, raw_discount, sale_price, discount_column_is_sale_price
    ):
        if sale_price is not None and price is not None:
            return max(price - sale_price, Decimal("0.00"))
        if raw_discount is None:
            return Decimal("0.00")
        if price is None or not discount_column_is_sale_price:
            return raw_discount
        return max(price - raw_discount, Decimal("0.00"))

    @staticmethod
    def detect_discount_column_as_sale_price(samples):
        if not samples:
            return False
        sale_like = 0
        for price, discount in samples:
            if price <= 0:
                continue
            if discount >= price or discount / price >= Decimal("0.7"):
                sale_like += 1
        return sale_like * 2 >= len(samples)

    @staticmethod
    def parse_datetime(value):
        if value is None or value == "":
            return None
        if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
            return value
        return None
