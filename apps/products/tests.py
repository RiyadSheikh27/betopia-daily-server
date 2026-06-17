import tempfile
from decimal import Decimal
from pathlib import Path

from django.core.management import call_command
from django.test import TestCase
from openpyxl import Workbook

from .models import Brand, Category, Product


class BulkUploadProductsCommandTest(TestCase):
    def test_bulk_upload_products_creates_brands_categories_and_products(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(
            [
                "Brand",
                "Category",
                "Subcategory",
                "Item Name",
                "Price",
                "Discount Amount",
                "SKU",
                "Unit",
                "Description",
                "In Stock",
            ]
        )
        sheet.append(
            [
                "TestBrand",
                "Food",
                "Food > Snacks",
                "Test Product",
                120.5,
                10,
                "TB-001",
                "per piece",
                "Sample description",
                "yes",
            ]
        )

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            workbook.save(tmp_file.name)
            tmp_path = Path(tmp_file.name)

        try:
            call_command("bulk_upload_products", str(tmp_path))

            self.assertTrue(Brand.objects.filter(name="TestBrand").exists())
            self.assertTrue(Category.objects.filter(name="Food").exists())
            self.assertTrue(Category.objects.filter(name="Food > Snacks").exists())
            self.assertTrue(
                Product.objects.filter(name="Test Product", sku="TB-001").exists()
            )
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_bulk_upload_reverses_price_and_discount_when_sale_price_is_in_discount_column(
        self,
    ):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sale Price Column"
        sheet.append(
            [
                "Brand",
                "Category",
                "SKU",
                "Item Name",
                "Brand Name",
                "unit",
                "price",
                "discount_amount",
            ]
        )
        sheet.append(
            [
                "Fortune",
                "Cooking",
                "RB-001",
                "Rice Bran Oil",
                "Fortune Fortified Rice Bran Oil",
                "Per ltr",
                1175,
                1170,
            ]
        )

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            workbook.save(tmp_file.name)
            tmp_path = Path(tmp_file.name)

        try:
            call_command("bulk_upload_products", str(tmp_path))
            product = Product.objects.get(sku="RB-001")
            self.assertEqual(product.price, Decimal("1175.00"))
            self.assertEqual(product.discount_amount, Decimal("5.00"))
            self.assertEqual(product.discounted_price, Decimal("1170.00"))
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_bulk_upload_with_hyphenated_sku_header(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Hyphen Header"
        sheet.append(
            [
                "Brand",
                "Category",
                "SKU-",
                "Item Name",
                "Brand Name",
                "unit",
                "price",
                "discount_amount",
            ]
        )
        sheet.append(
            [
                "HyphenBrand",
                "Snacks",
                "HY-001",
                "Hyphen Test Product",
                "HyphenBrand",
                "per piece",
                100,
                10,
            ]
        )

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            workbook.save(tmp_file.name)
            tmp_path = Path(tmp_file.name)

        try:
            call_command("bulk_upload_products", str(tmp_path))

            self.assertTrue(Brand.objects.filter(name="HyphenBrand").exists())
            self.assertTrue(Category.objects.filter(name="Snacks").exists())
            self.assertTrue(
                Product.objects.filter(
                    name="Hyphen Test Product", sku="HY-001"
                ).exists()
            )
        finally:
            tmp_path.unlink(missing_ok=True)
